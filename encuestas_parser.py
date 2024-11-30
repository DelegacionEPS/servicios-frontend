import openpyxl

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("encuestas.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(1, dataframe1.max_row):
    data = []
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        data.append(col[row].value)

    percentage = str(data[1])[2:]

    if (len(percentage) < 3):
        percentage = percentage + ".00%"
    else:
        percentage = percentage[0:2] + "." + percentage[2:] + "%"
    print(str(data[0]) + "," + percentage)
